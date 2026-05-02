from io import BytesIO

from openpyxl import Workbook, load_workbook


def test_dashboard_requires_auth(client):
    response = client.get("/api/admin/dashboard")
    assert response.status_code == 401


def test_category_and_product_crud(client, admin_cookie):
    method_response = client.post(
        "/api/admin/methods",
        json={
            "name": "Custom HPLC",
            "slug": "custom-hplc",
            "description": "Custom HPLC method for marker assay.",
            "sort_order": 20,
            "is_active": True,
        },
        cookies=admin_cookie,
    )
    assert method_response.status_code == 201
    method = method_response.json()

    category_response = client.post(
        "/api/admin/categories",
        json={
            "name": "Botanical Powders",
            "slug": "botanical-powders",
            "description": "Spray-dried botanicals",
            "sort_order": 10,
            "is_active": True,
        },
        cookies=admin_cookie,
    )
    assert category_response.status_code == 201
    category = category_response.json()

    product_response = client.post(
        "/api/admin/products",
        json={
            "category_id": category["id"],
            "common_name": "Moringa Powder",
            "botanical_name": "Moringa oleifera",
            "specification": "80 mesh, low moisture",
            "method_ids": [method["id"]],
            "sort_order": 1,
            "is_active": True,
        },
        cookies=admin_cookie,
    )
    assert product_response.status_code == 201
    product = product_response.json()
    assert product["category"]["slug"] == "botanical-powders"
    assert product["methods"][0]["slug"] == "custom-hplc"

    update_response = client.put(
        f"/api/admin/products/{product['id']}",
        json={"specification": "100 mesh, low moisture", "method_ids": []},
        cookies=admin_cookie,
    )
    assert update_response.status_code == 200
    assert update_response.json()["specification"] == "100 mesh, low moisture"
    assert update_response.json()["methods"] == []


def test_method_crud(client, admin_cookie):
    create_response = client.post(
        "/api/admin/methods",
        json={
            "name": "Test LC-MS/MS",
            "slug": "test-lc-ms-ms",
            "description": "Tandem mass spectrometry method for testing.",
            "sort_order": 15,
            "is_active": True,
        },
        cookies=admin_cookie,
    )
    assert create_response.status_code == 201
    created = create_response.json()

    list_response = client.get("/api/admin/methods", cookies=admin_cookie)
    assert list_response.status_code == 200
    assert any(item["slug"] == "test-lc-ms-ms" for item in list_response.json())

    update_response = client.put(
        f"/api/admin/methods/{created['id']}",
        json={"description": "Updated method description"},
        cookies=admin_cookie,
    )
    assert update_response.status_code == 200
    assert update_response.json()["description"] == "Updated method description"

    delete_response = client.delete(f"/api/admin/methods/{created['id']}", cookies=admin_cookie)
    assert delete_response.status_code == 204


def test_product_excel_template_and_upload(client, admin_cookie):
    template_response = client.get("/api/admin/products/template", cookies=admin_cookie)
    assert template_response.status_code == 200
    workbook = load_workbook(BytesIO(template_response.content))
    assert workbook.sheetnames[0] == "Products"
    assert [cell.value for cell in workbook["Products"][1]] == [
        "category_slug",
        "common_name",
        "botanical_name",
        "specification",
        "methods",
        "sort_order",
        "is_active",
    ]
    assert workbook["Products"]["A2"].value
    assert "Reference" in workbook.sheetnames

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append([
        "category_slug",
        "common_name",
        "botanical_name",
        "specification",
        "methods",
        "sort_order",
        "is_active",
    ])
    sheet.append([
        "herbal-extracts",
        "Excel Product",
        "Excel botanica",
        "Marker 10%",
        "HPLC, Microbiological Testing",
        99,
        "TRUE",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    upload_response = client.post(
        "/api/admin/products/upload",
        files={
            "file": (
                "products.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        cookies=admin_cookie,
    )
    assert upload_response.status_code == 200
    assert upload_response.json()["created"] == 1

    products_response = client.get("/api/admin/products", cookies=admin_cookie)
    product = next(item for item in products_response.json() if item["common_name"] == "Excel Product")
    assert product["sort_order"] == 99
    assert [method["slug"] for method in product["methods"]] == [
        "hplc",
        "microbiological-testing",
    ]


def test_inquiry_status_update(client, admin_cookie):
    create_response = client.post(
        "/api/inquiries",
        json={
            "source": "inquiry",
            "name": "Aman Verma",
            "company_name": "Nutra World",
            "email": "aman@example.com",
            "phone": "+91 98111 11111",
            "product_requirement": "Vitamin C",
            "message": "Need pricing for export shipment to UAE.",
        },
    )
    assert create_response.status_code == 201
    inquiry_id = create_response.json()["inquiry"]["id"]

    update_response = client.patch(
        f"/api/admin/inquiries/{inquiry_id}",
        json={"status": "contacted"},
        cookies=admin_cookie,
    )
    assert update_response.status_code == 200
    assert update_response.json()["status"] == "contacted"
