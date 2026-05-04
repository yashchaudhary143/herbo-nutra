from datetime import time
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_current_admin
from app.db.session import get_db
from app.models import Category, Inquiry, Method, Product
from app.schemas.category import CategoryCreate, CategoryRead, CategoryUpdate
from app.schemas.dashboard import DashboardStats
from app.schemas.inquiry import InquiryRead, InquiryUpdate, PaginatedInquiries
from app.schemas.method import MethodCreate, MethodRead, MethodUpdate
from app.schemas.product import ProductCreate, ProductRead, ProductUpdate

router = APIRouter(prefix="/admin", tags=["admin"])

PRODUCT_IMPORT_HEADERS = [
    "category_slug",
    "common_name",
    "botanical_name",
    "specification",
    "methods",
    "sort_order",
    "is_active",
]

METHOD_IMPORT_ALIASES = {
    "by hplc": "hplc",
    "by uv": "uv-vis",
    "gc": "gc-ms",
    "gv": "gravimetry",
    "uv": "uv-vis",
}


def _normalize_import_value(value: object) -> str:
    if isinstance(value, time):
        return f"{value.hour}:{value.minute}"
    return str(value).strip().lstrip("\ufeff") if value is not None else ""


def _normalize_import_cell_value(cell: object) -> str:
    value = getattr(cell, "value", None)
    number_format = str(getattr(cell, "number_format", ""))
    if isinstance(value, (int, float)) and "%" in number_format:
        percentage = value * 100
        if percentage.is_integer():
            return f"{int(percentage)}%"
        return f"{percentage:g}%"
    return _normalize_import_value(value)


def _normalize_lookup(value: str) -> str:
    return value.strip().lower()


def _parse_bool(value: object, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1", "active"}:
        return True
    if normalized in {"false", "no", "n", "0", "inactive"}:
        return False
    raise ValueError("Use TRUE/FALSE, YES/NO, 1/0, Active/Inactive, or leave blank")


def _parse_methods(value: object) -> list[str]:
    raw = _normalize_import_value(value)
    if not raw:
        return []
    normalized = raw.replace("\n", ",").replace(";", ",").replace("&", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def _normalize_method_lookup_key(value: str) -> str:
    key = _normalize_lookup(value)
    return METHOD_IMPORT_ALIASES.get(key, key)


def _build_product_template(categories: list[Category], methods: list[Method]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    workbook.active = 0

    sample_row = [
        categories[0].slug if categories else "herbal-extracts",
        "Ashwagandha Extract",
        "Withania somnifera",
        "Withanolides 5%",
        "HPLC, HPTLC, Microbiological Testing",
        1,
        "TRUE",
    ]

    for column_number, header in enumerate(PRODUCT_IMPORT_HEADERS, start=1):
        sheet.cell(row=1, column=column_number, value=header)
    for column_number, value in enumerate(sample_row, start=1):
        sheet.cell(row=2, column=column_number, value=value)

    header_fill = PatternFill("solid", fgColor="1F5937")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    for column, width in {
        "A": 34,
        "B": 28,
        "C": 28,
        "D": 42,
        "E": 46,
        "F": 14,
        "G": 14,
    }.items():
        sheet.column_dimensions[column].width = width

    reference = workbook.create_sheet("Reference")
    reference.append(["Categories: use category_slug in Products sheet"])
    for category in categories:
        reference.append([category.slug, category.name])
    reference.append([])
    reference.append(["Methods: use names or slugs, comma-separated"])
    for method in methods:
        reference.append([method.slug, method.name])
    reference.column_dimensions["A"].width = 34
    reference.column_dimensions["B"].width = 52

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _import_products_from_workbook(file_bytes: bytes, db: Session) -> dict[str, object]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Upload a valid .xlsx workbook") from exc

    sheet = workbook["Products"] if "Products" in workbook.sheetnames else workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_normalize_import_value(value).lower() for value in (header_row or [])]
    missing_headers = [header for header in PRODUCT_IMPORT_HEADERS if header not in headers]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing_headers)}")

    column_index = {header: headers.index(header) for header in PRODUCT_IMPORT_HEADERS}
    categories = db.query(Category).all()
    methods = db.query(Method).all()
    category_lookup = {
        **{_normalize_lookup(category.slug): category for category in categories},
        **{_normalize_lookup(category.name): category for category in categories},
    }
    method_lookup = {
        **{_normalize_lookup(method.slug): method for method in methods},
        **{_normalize_lookup(method.name): method for method in methods},
    }

    rows: list[dict[str, object]] = []
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if not any(_normalize_import_cell_value(cell) for cell in row):
            continue

        category_key = _normalize_lookup(_normalize_import_cell_value(row[column_index["category_slug"]]))
        category = category_lookup.get(category_key)
        common_name = _normalize_import_cell_value(row[column_index["common_name"]])
        botanical_name = _normalize_import_cell_value(row[column_index["botanical_name"]])
        specification = _normalize_import_cell_value(row[column_index["specification"]])

        if not category:
            errors.append(f"Row {row_number}: unknown category '{category_key or 'blank'}'")
        if not common_name:
            errors.append(f"Row {row_number}: common_name is required")
        if not botanical_name:
            errors.append(f"Row {row_number}: botanical_name is required")
        if not specification:
            errors.append(f"Row {row_number}: specification is required")

        row_methods: list[Method] = []
        for method_name in _parse_methods(row[column_index["methods"]].value):
            method = method_lookup.get(_normalize_method_lookup_key(method_name))
            if not method:
                errors.append(f"Row {row_number}: unknown method '{method_name}'")
                continue
            if method not in row_methods:
                row_methods.append(method)

        try:
            sort_order_raw = row[column_index["sort_order"]].value
            sort_order = int(sort_order_raw) if _normalize_import_value(sort_order_raw) else 0
        except (TypeError, ValueError):
            errors.append(f"Row {row_number}: sort_order must be a number")
            sort_order = 0

        try:
            is_active = _parse_bool(row[column_index["is_active"]].value)
        except ValueError as exc:
            errors.append(f"Row {row_number}: is_active {exc}")
            is_active = True

        if category and common_name and botanical_name and specification:
            rows.append(
                {
                    "category": category,
                    "common_name": common_name,
                    "botanical_name": botanical_name,
                    "specification": specification,
                    "methods": row_methods,
                    "sort_order": sort_order,
                    "is_active": is_active,
                }
            )

    if errors:
        raise HTTPException(status_code=400, detail={"message": "Product import failed", "errors": errors})
    if not rows:
        raise HTTPException(status_code=400, detail="No product rows found in workbook")

    created = 0
    updated = 0
    for row in rows:
        product = db.query(Product).filter(
            Product.common_name == row["common_name"],
            Product.botanical_name == row["botanical_name"],
        ).first()
        if product:
            updated += 1
        else:
            product = Product(
                common_name=str(row["common_name"]),
                botanical_name=str(row["botanical_name"]),
            )
            created += 1

        product.category_id = row["category"].id
        product.specification = str(row["specification"])
        product.sort_order = int(row["sort_order"])
        product.is_active = bool(row["is_active"])
        product.methods = row["methods"]
        db.add(product)

    db.commit()
    return {"created": created, "updated": updated, "total": len(rows)}


@router.get("/dashboard", response_model=DashboardStats)
def dashboard(_: object = Depends(get_current_admin), db: Session = Depends(get_db)) -> DashboardStats:
    return DashboardStats(
        categories=db.query(func.count(Category.id)).scalar() or 0,
        products=db.query(func.count(Product.id)).scalar() or 0,
        inquiries=db.query(func.count(Inquiry.id)).scalar() or 0,
        new_inquiries=db.query(func.count(Inquiry.id)).filter(Inquiry.status == "new").scalar() or 0,
    )


@router.get("/methods", response_model=list[MethodRead])
def admin_methods(_: object = Depends(get_current_admin), db: Session = Depends(get_db)) -> list[MethodRead]:
    items = db.query(Method).order_by(Method.sort_order.asc(), Method.name.asc()).all()
    return [MethodRead.model_validate(item) for item in items]


@router.post("/methods", response_model=MethodRead, status_code=status.HTTP_201_CREATED)
def create_method(
    payload: MethodCreate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> MethodRead:
    existing = db.query(Method).filter((Method.slug == payload.slug) | (Method.name == payload.name)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Method name or slug already exists")

    item = Method(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return MethodRead.model_validate(item)


@router.put("/methods/{method_id}", response_model=MethodRead)
def update_method(
    method_id: int,
    payload: MethodUpdate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> MethodRead:
    item = db.get(Method, method_id)
    if not item:
        raise HTTPException(status_code=404, detail="Method not found")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.add(item)
    db.commit()
    db.refresh(item)
    return MethodRead.model_validate(item)


@router.delete("/methods/{method_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_method(
    method_id: int,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> None:
    item = db.get(Method, method_id)
    if not item:
        raise HTTPException(status_code=404, detail="Method not found")
    db.delete(item)
    db.commit()


@router.get("/categories", response_model=list[CategoryRead])
def admin_categories(_: object = Depends(get_current_admin), db: Session = Depends(get_db)) -> list[CategoryRead]:
    rows = (
        db.query(Category, func.count(Product.id).label("product_count"))
        .outerjoin(Product, Product.category_id == Category.id)
        .group_by(Category.id)
        .order_by(Category.sort_order.asc(), Category.name.asc())
        .all()
    )
    return [
        CategoryRead.model_validate({**category.__dict__, "product_count": product_count})
        for category, product_count in rows
    ]


@router.post("/categories", response_model=CategoryRead, status_code=status.HTTP_201_CREATED)
def create_category(
    payload: CategoryCreate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CategoryRead:
    existing = db.query(Category).filter((Category.slug == payload.slug) | (Category.name == payload.name)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Category name or slug already exists")

    category = Category(**payload.model_dump())
    db.add(category)
    db.commit()
    db.refresh(category)
    return CategoryRead.model_validate({**category.__dict__, "product_count": 0})


@router.put("/categories/{category_id}", response_model=CategoryRead)
def update_category(
    category_id: int,
    payload: CategoryUpdate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CategoryRead:
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(category, key, value)
    db.add(category)
    db.commit()
    db.refresh(category)
    product_count = db.query(func.count(Product.id)).filter(Product.category_id == category.id).scalar() or 0
    return CategoryRead.model_validate({**category.__dict__, "product_count": product_count})


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_category(
    category_id: int,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> None:
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    db.delete(category)
    db.commit()


@router.get("/products", response_model=list[ProductRead])
def admin_products(
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> list[ProductRead]:
    items = (
        db.query(Product)
        .options(joinedload(Product.category), joinedload(Product.methods))
        .order_by(Product.sort_order.asc(), Product.common_name.asc())
        .all()
    )
    return [ProductRead.model_validate(item) for item in items]


@router.get("/products/template")
def product_import_template(
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    categories = db.query(Category).order_by(Category.sort_order.asc(), Category.name.asc()).all()
    methods = db.query(Method).order_by(Method.sort_order.asc(), Method.name.asc()).all()
    buffer = _build_product_template(categories, methods)
    headers = {"Content-Disposition": 'attachment; filename="product-upload-template.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/products/upload")
async def upload_products(
    file: UploadFile = File(...),
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file")

    result = _import_products_from_workbook(await file.read(), db)
    return {"message": "Products imported successfully", **result}


@router.post("/products", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
def create_product(
    payload: ProductCreate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> ProductRead:
    category = db.get(Category, payload.category_id)
    if not category:
        raise HTTPException(status_code=400, detail="Category not found")

    method_ids = payload.method_ids
    methods = db.query(Method).filter(Method.id.in_(method_ids)).all() if method_ids else []
    if method_ids and len(methods) != len(set(method_ids)):
        raise HTTPException(status_code=400, detail="One or more methods were not found")

    product = Product(**payload.model_dump(exclude={"method_ids"}))
    product.methods = methods
    db.add(product)
    db.commit()
    db.refresh(product)
    product = (
        db.query(Product)
        .options(joinedload(Product.category), joinedload(Product.methods))
        .filter(Product.id == product.id)
        .first()
    )
    return ProductRead.model_validate(product)


@router.put("/products/{product_id}", response_model=ProductRead)
def update_product(
    product_id: int,
    payload: ProductUpdate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> ProductRead:
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    updates = payload.model_dump(exclude_unset=True)
    if "category_id" in updates and not db.get(Category, updates["category_id"]):
        raise HTTPException(status_code=400, detail="Category not found")
    if "method_ids" in updates:
        method_ids = updates.pop("method_ids") or []
        methods = db.query(Method).filter(Method.id.in_(method_ids)).all() if method_ids else []
        if len(methods) != len(set(method_ids)):
            raise HTTPException(status_code=400, detail="One or more methods were not found")
        product.methods = methods

    for key, value in updates.items():
        setattr(product, key, value)
    db.add(product)
    db.commit()
    refreshed = (
        db.query(Product)
        .options(joinedload(Product.category), joinedload(Product.methods))
        .filter(Product.id == product.id)
        .first()
    )
    return ProductRead.model_validate(refreshed)


@router.delete("/products/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(
    product_id: int,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> None:
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(product)
    db.commit()


@router.get("/inquiries", response_model=PaginatedInquiries)
def list_inquiries(
    status_filter: str | None = Query(default=None, alias="status"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=20, ge=1, le=100),
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> PaginatedInquiries:
    query = db.query(Inquiry)
    if status_filter:
        query = query.filter(Inquiry.status == status_filter)
    total = query.count()
    items = (
        query.order_by(Inquiry.created_at.desc())
        .offset((page - 1) * limit)
        .limit(limit)
        .all()
    )
    return PaginatedInquiries(
        items=[InquiryRead.model_validate(item) for item in items],
        total=total,
        page=page,
        limit=limit,
    )


@router.patch("/inquiries/{inquiry_id}", response_model=InquiryRead)
def update_inquiry(
    inquiry_id: int,
    payload: InquiryUpdate,
    _: object = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> InquiryRead:
    inquiry = db.get(Inquiry, inquiry_id)
    if not inquiry:
        raise HTTPException(status_code=404, detail="Inquiry not found")
    inquiry.status = payload.status
    db.add(inquiry)
    db.commit()
    db.refresh(inquiry)
    return InquiryRead.model_validate(inquiry)
